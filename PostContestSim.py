#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UFC Post-Contest Simulator — Multi-Contest (shared fight sampling) + HP + MP

v3 performance release — byte-identical outputs, ~5x faster engine:
  * Ranking rewritten: one argsort + tie-run boundary scan replaces the flat
    double-searchsorted (which was 30x the cost of the sort itself at big
    contest sizes). Rank intervals come out bit-identical.
  * Deduplication: entries with identical mapped fighter columns are one
    unique lineup; all rank/payout math runs on uniques with copy weights
    (~2-3x fewer columns on big GPPs). A one-time per-contest self-check
    bit-compares unique vs entry scores and silently falls back to the legacy
    path if a BLAS/numpy change ever breaks the equivalence. Degenerate
    shapes (1-row micro-batch tails, 1-unique contests) always use legacy.
  * Wins/2nd/3rd derived from the rank intervals (no extra masked-max passes).
  * Portfolio tracking stores per-contest outcomes only for users IN each
    contest (ragged) instead of (contests x iters x ALL users) dense — ~7x
    less RAM/disk; the dense arrays were the real ceiling on iteration count.
  * Percentile phase sorts all users as bulk matrix ops (was: one Python-loop
    sort per user per contest); result CSVs built column-wise (was: per-row).
  * Worker results are accumulated in submission order, so a fixed seed now
    reproduces byte-identical outputs run to run.
  * BLAS pinned to 1 thread per worker (workers are the parallelism); the
    workers prompt now suggests your machine's core count.
  * --legacy flag forces the original ranking path everywhere.

v2 performance/stability release — identical outputs, faster engine:
  * Work is now split across ALL workers regardless of progress step.
    (v1 made one worker task per progress_step; with defaults 200k iters /
    100k step only 2 of 4 workers ever ran.)
  * Payout engine fully vectorized across the micro-batch: one sort +
    two flat searchsorted calls per batch instead of a Python loop with
    ~10 numpy calls per iteration per contest.
  * Per-user payout aggregation via precomputed group-reduce
    (np.add.reduceat) instead of one bincount per iteration.
  * Optimal-lineup tally: iterations where the plain top-6 scorers fit
    the salary cap (the common case) skip the brute-force combo matmul.
  * WinPct/SecondPct/ThirdPct semantics unchanged: ties at the top all
    count as wins (matches DK displayed rank). Payout tie-splitting
    unchanged (dead-heat).
  * Fixed: double "Press Enter to close" prompt; lineup-sheet padding
    crash when a header cell isn't text.

Behavior:
- Multi-contest: contests defined on workbook sheet 'Contests'
    Required columns (case-insensitive):
      ContestName, LineupsSheet, PayoutsSheet, EntryFee
    Optional:
      PrefixSheet   (cumulative payout prefix; validated; otherwise computed from payouts)
- LineupsSheet can be either:
    * An Excel sheet name (original behavior)
    * A DraftKings CSV file path ending in .csv (e.g. "contest1.csv")
      CSV files are resolved relative to the workbook directory.
      Expected CSV columns: Rank, EntryId, EntryName, TimeRemaining, Points, Lineup
- Fight score sampling is done ONCE per iteration and reused across contests.
- Missing/empty fight sim sheet(s) => fail fast.
- Fighter name normalization (whitespace collapse). Unmapped fighters score 0.
- Blank lineup rows with a username are kept and will score 0.
- Outputs per-contest results CSVs (one file per contest). No portfolio summary, no series.
- Sanity check per contest: avg EV equals prize_pool/n and avg total paid equals prize_pool.
Backwards compatible fallback:
- If 'Contests' sheet is missing, runs single contest:
    ContestName='Main', LineupsSheet='Post Contest Sim', PayoutsSheet='Payouts',
    PrefixSheet='PayoutPrefix' (if present), EntryFee read from 'DraftKings Fighter Pool'!D2.
"""
import os, sys, math, time, argparse, datetime, re, tempfile, uuid, json
from pathlib import Path
from typing import Optional
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from concurrent.futures import ProcessPoolExecutor
import itertools
Z99 = 2.326347874
# Optimal-lineup (DraftKings classic MMA): 6 fighters, $50,000 cap, no pairing rule.
OPT_ROSTER = 6
OPT_CAP = 50000
_money_re = re.compile(r'[^0-9.\-]')
_nbsp = '\xa0'
import atexit
_already_paused = False
def _pause_on_exit():
    if _already_paused:
        return
    try:
        if sys.stdin and sys.stdin.isatty():
            input("Press Enter to close...")
    except Exception:
        pass
atexit.register(_pause_on_exit)
# -------------------- Utilities --------------------
def to_money(x):
    if x is None: return 0.0
    if isinstance(x, float) and math.isnan(x): return 0.0
    try: s = str(x)
    except Exception: return 0.0
    s = _money_re.sub('', s)
    if s in ('', '-', '.', '-.'): return 0.0
    try: return float(s)
    except Exception: return 0.0
def safe_str(x) -> str:
    if x is None: return ""
    if isinstance(x, float) and math.isnan(x): return ""
    try: s = str(x).strip()
    except Exception: return ""
    if s.lower() in ("nan","<na>","none"): return ""
    return s
def norm_name_fighter(s: str) -> str:
    s = safe_str(s).replace(_nbsp, ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s
def norm_user(u: str) -> str:
    s = safe_str(u).replace(_nbsp, ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s.casefold()
def log(msg): print(msg, flush=True)
def safe_filename(s: str) -> str:
    s = safe_str(s)
    s = re.sub(r'[^A-Za-z0-9 _\-]+', '', s)
    s = s.strip().replace(' ', '_')
    return s if s else "Contest"
# -------------------- DraftKings CSV helpers --------------------
def strip_entry_number(username: str) -> str:
    """Strip DraftKings entry number suffix like ' (2)' or ' (1/9)' from username."""
    return re.sub(r'\s*\(\d+(?:/\d+)?\)\s*$', '', safe_str(username).strip())

def parse_dk_lineup_string(s: str) -> list:
    """Parse a DraftKings lineup string into individual fighter names.

    Input format:  'F Anthony Hernandez F Melquizael Costa F Zach Reese ...'
    Output:        ['Anthony Hernandez', 'Melquizael Costa', 'Zach Reese', ...]
    """
    s = safe_str(s).strip()
    if not s:
        return []
    parts = re.split(r'\s+F\s+', s)
    if parts and parts[0].startswith('F '):
        parts[0] = parts[0][2:]
    fighters = [p.strip() for p in parts if p.strip()]
    return fighters

def read_lineups_csv(csv_path: str):
    """Read lineups from a DraftKings contest export CSV file.

    Expected columns: Rank, EntryId, EntryName, TimeRemaining, Points, Lineup
    Returns (fighters_array, users_list).
    """
    if not Path(csv_path).exists():
        raise FileNotFoundError(f"DraftKings CSV file not found: {csv_path}")

    if csv_path.lower().endswith('.zip'):
        # DK standings download is a .zip holding one CSV — read it in place so
        # there's no manual unzip step.
        import zipfile
        with zipfile.ZipFile(csv_path) as z:
            inner = [n for n in z.namelist() if n.lower().endswith('.csv')]
            if not inner:
                raise ValueError(f"No CSV found inside zip: {csv_path}")
            with z.open(inner[0]) as fh:
                df = pd.read_csv(fh)
    else:
        df = pd.read_csv(csv_path)

    col_map = {c.strip().lower(): c for c in df.columns}
    entry_col = col_map.get('entryname')
    lineup_col = col_map.get('lineup')

    if entry_col is None:
        raise ValueError(f"CSV file '{csv_path}' missing 'EntryName' column. "
                         f"Found columns: {list(df.columns)}")
    if lineup_col is None:
        raise ValueError(f"CSV file '{csv_path}' missing 'Lineup' column. "
                         f"Found columns: {list(df.columns)}")

    fighters_list = []
    users = []

    for _, row in df.iterrows():
        username = strip_entry_number(safe_str(row[entry_col]))
        lineup_str = safe_str(row[lineup_col])
        parsed = parse_dk_lineup_string(lineup_str)

        while len(parsed) < 6:
            parsed.append("")
        parsed = parsed[:6]

        fighters_list.append(parsed)
        users.append(username)

    fighters = np.array(fighters_list, dtype=object)

    mask = np.array([(any(bool(fighters[i, c]) for c in range(6)) or bool(users[i]))
                     for i in range(len(fighters))])

    log(f"[csv] loaded {mask.sum():,} entries from {Path(csv_path).name}")
    return fighters[mask], [u for i, u in enumerate(users) if mask[i]]

# -------------------- Workbook readers --------------------
def read_entry_fee_fallback(wb_path: str) -> float:
    """Fallback location: DraftKings Fighter Pool!D2."""
    wb = load_workbook(wb_path, data_only=True, read_only=True)
    if "DraftKings Fighter Pool" not in wb.sheetnames:
        raise FileNotFoundError("Sheet 'DraftKings Fighter Pool' not found.")
    ws = wb["DraftKings Fighter Pool"]
    fee = to_money(ws["D2"].value)
    if fee <= 0:
        raise ValueError("Entry fee missing/invalid at DraftKings Fighter Pool!D2.")
    return float(fee)
def read_prefix_sheet_named(xl: pd.ExcelFile, prefix_sheet: Optional[str]):
    if not prefix_sheet:
        return None
    try:
        return pd.read_excel(xl, sheet_name=prefix_sheet, engine="openpyxl").iloc[:, :2]
    except Exception:
        return None
def read_payouts_named(xl: pd.ExcelFile, payouts_sheet: str, prefix_sheet: Optional[str]):
    # Per-rank payouts
    df = pd.read_excel(xl, sheet_name=payouts_sheet, engine="openpyxl").iloc[:, :2].copy()
    df = df.dropna(how="all")
    if df.shape[1] < 2:
        raise ValueError(f"Payouts sheet '{payouts_sheet}' must have at least 2 columns (Rank, Payout).")
    rank_col, money_col = df.columns[:2]
    df[rank_col]  = pd.to_numeric(df[rank_col], errors="coerce")
    df[money_col] = pd.to_numeric(df[money_col].map(to_money), errors="coerce").astype("float64")
    df = df.dropna(subset=[rank_col]).sort_values(rank_col)
    if df.empty:
        raise ValueError(f"Payouts sheet '{payouts_sheet}' is empty/invalid.")
    max_rank = int(df[rank_col].max())
    payouts = [0.0] * (max_rank + 1)
    for _, row in df.iterrows():
        r = int(row[rank_col])
        if 1 <= r <= max_rank:
            payouts[r] = float(row[money_col])
    last_paid = max([i for i, v in enumerate(payouts) if v > 0], default=0)
    sum_payouts = float(np.nansum(payouts))
    # Optional cumulative prefix sheet; validate it
    prefix = None
    dfp = read_prefix_sheet_named(xl, prefix_sheet)
    if dfp is not None:
        try:
            if dfp.shape[1] < 2:
                raise ValueError("Prefix sheet must have 2 columns (Rank, CumulativePayout).")
            rcol, pcol = dfp.columns[:2]
            dfp = dfp[[rcol, pcol]].copy()
            dfp[rcol] = pd.to_numeric(dfp[rcol], errors="coerce")
            dfp[pcol] = pd.to_numeric(dfp[pcol].map(to_money), errors="coerce").astype("float64")
            dfp = dfp.dropna(subset=[rcol]).sort_values(rcol)
            max_rank2 = int(dfp[rcol].max())
            M = max(max_rank, max_rank2)
            if M > max_rank:
                payouts += [0.0] * (M - max_rank)
                max_rank = M
            prefix = [0.0] * (max_rank + 1)
            for _, row in dfp.iterrows():
                r = int(row[rcol])
                if 1 <= r <= max_rank:
                    prefix[r] = float(row[pcol])
            # Forward-fill missing cumulative by adding per-rank payouts
            for r in range(1, max_rank + 1):
                if prefix[r] == 0.0:
                    prefix[r] = prefix[r-1] + (payouts[r] if r < len(payouts) else 0.0)
            if last_paid < len(prefix) and abs(prefix[last_paid] - sum_payouts) > 1e-6:
                log(f"[warn] Prefix sheet '{prefix_sheet}' doesn't match per-rank payouts in '{payouts_sheet}'; recomputing prefix from payouts.")
                prefix = None
        except Exception:
            prefix = None
    if prefix is None:
        prefix = [0.0] * len(payouts)
        for r in range(1, len(payouts)):
            prefix[r] = prefix[r-1] + payouts[r]
    return np.array(payouts, dtype=np.float64), np.array(prefix, dtype=np.float64), int(last_paid)

def read_payouts_json(ref: str, wb_path: str):
    """Load a per-rank payout table from a scraped dk_payouts file instead of an
    Excel sheet, so payout tables never have to be entered by hand. `ref` is a
    DraftKings contest id (all digits) -> dk_payouts/<id>.json next to the
    workbook, or a direct path to such a JSON. Returns (payouts, prefix,
    last_paid) in the exact shape read_payouts_named returns: arrays indexed by
    rank (index 0 unused), prefix cumulative."""
    s = str(ref).strip()
    base = Path(wb_path).resolve().parent
    if s.isdigit():
        path = base / "dk_payouts" / f"{s}.json"
    else:
        p = Path(s)
        path = p if p.is_absolute() else (base / s)
    if not path.exists():
        hint = s if s.isdigit() else "<contest-id>"
        raise FileNotFoundError(
            f"Scraped payout file not found: {path}\n"
            f"  Build it with: python scrape_dk_contests.py --payouts --ids {hint}")
    with open(path, "r", encoding="utf-8") as f:
        rec = json.load(f)
    per_rank = rec.get("payouts") or []
    if not per_rank:
        raise ValueError(f"Payout file '{path}' has no non-empty 'payouts' list.")
    max_rank = len(per_rank)
    payouts = [0.0] * (max_rank + 1)            # payouts[r] = prize for rank r
    for i, v in enumerate(per_rank):
        payouts[i + 1] = float(v)
    prefix = [0.0] * (max_rank + 1)             # prefix[r] = cumulative through rank r
    for r in range(1, max_rank + 1):
        prefix[r] = prefix[r - 1] + payouts[r]
    last_paid = max([r for r in range(1, max_rank + 1) if payouts[r] > 0], default=0)
    return np.array(payouts, dtype=np.float64), np.array(prefix, dtype=np.float64), int(last_paid)

def load_payouts(xl, ref: str, prefix_sheet, wb_path: str):
    """Dispatch payout loading: a numeric DraftKings id or a *.json path pulls
    from a scraped dk_payouts file; anything else is an Excel sheet name (the
    original behavior, kept as a fallback so existing workbooks still run)."""
    s = str(ref).strip()
    if s.isdigit() or s.lower().endswith(".json"):
        return read_payouts_json(s, wb_path)
    return read_payouts_named(xl, ref, prefix_sheet)

def read_fighter_map(xl: pd.ExcelFile):
    # Columns: Fighter (A), FightID (B), optional Score (C), optional Salary (D)
    df = pd.read_excel(xl, sheet_name="DraftKings Fighter Pool", engine="openpyxl")
    ncols = df.shape[1]
    if ncols >= 4:
        df = df.iloc[:, :4]
        df.columns = ["Fighter", "FightID", "Score", "Salary"]
    elif ncols >= 3:
        df = df.iloc[:, :3]
        df.columns = ["Fighter", "FightID", "Score"]
        df["Salary"] = np.nan
    else:
        df = df.iloc[:, :2]
        df.columns = ["Fighter", "FightID"]
        df["Score"] = np.nan
        df["Salary"] = np.nan
    keep = ~(df["Fighter"].isna() & df["FightID"].isna())
    df = df.loc[keep].reset_index(drop=True)
    seen = {}
    fmap = {}
    fixed_scores = {}  # fighter_name -> fixed DK score (float)
    salary_map = {}    # fighter_name -> salary (float)
    fighter_order = []  # ordered list of (name, fid, slot) for fight card
    for _, row in df.iterrows():
        name = norm_name_fighter(row["Fighter"])
        if not name:
            continue
        try:
            fid = int(pd.to_numeric(row["FightID"], errors="coerce"))
        except Exception:
            continue
        # Order in the pool defines slot 1 then slot 2 for a FightID
        if fid not in seen:
            seen[fid] = 1
            slot = 1
        else:
            seen[fid] += 1
            slot = 2
        fmap[name] = (fid, slot)
        score_val = pd.to_numeric(row["Score"], errors="coerce")
        if not (score_val is None or (isinstance(score_val, float) and math.isnan(score_val))):
            fixed_scores[name] = float(score_val)
        sal_val = pd.to_numeric(row["Salary"], errors="coerce")
        if not (sal_val is None or (isinstance(sal_val, float) and math.isnan(sal_val))):
            salary_map[name] = float(sal_val)
        fighter_order.append({"name": name, "fight_id": fid, "slot": slot})
    return fmap, fixed_scores, fighter_order, salary_map
def read_lineups_sheet(xl: pd.ExcelFile, sheet_name: str):
    # columns A:G => F1..F6 + Username
    df = pd.read_excel(xl, sheet_name=sheet_name, engine="openpyxl", usecols="A:G")
    if df.shape[1] < 7:
        for pad_i in range(7 - df.shape[1]):
            df[f"_pad{pad_i}"] = ""
    df = df.iloc[:, :7]
    fighters = np.empty((len(df), 6), dtype=object)
    users = []
    for i in range(len(df)):
        row = df.iloc[i]
        for c in range(6):
            fighters[i, c] = safe_str(row.iloc[c])
        users.append(strip_entry_number(safe_str(row.iloc[6])))
    # Keep rows if they have any fighter OR a username (blank lineup with username stays; scores 0)
    mask = np.array([(any(bool(fighters[i, c]) for c in range(6)) or bool(users[i]))
                     for i in range(len(df))])
    return fighters[mask], [u for i, u in enumerate(users) if mask[i]]

def read_lineups(xl: pd.ExcelFile, lineups_ref: str, wb_path: str):
    """Read lineups from a CSV/zip file (relative to the workbook) or an Excel sheet."""
    low = lineups_ref.lower()
    if low.endswith('.csv') or low.endswith('.zip'):
        csv_path = Path(wb_path).resolve().parent / lineups_ref
        return read_lineups_csv(str(csv_path))
    else:
        return read_lineups_sheet(xl, lineups_ref)

def _load_dk_lobby(base: Path):
    """id (str) -> lobby record, from dk_lobby.json next to the workbook (or {})."""
    p = base / "dk_lobby.json"
    if not p.exists():
        return {}
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return {}
    return {str(c.get("id")): c for c in data if c.get("id") is not None}

def _tidy_contest_name(dk_name: str, cid: str) -> str:
    """DK's verbose lobby name -> a clean display name matching the site's style.
    'UFC $26K Double Champ [$13K to 1st]' -> '26K Double Champ'."""
    if not dk_name:
        return f"Contest {cid}"
    s = re.sub(r'^\s*UFC\s+', '', dk_name)      # drop the 'UFC ' prefix
    s = re.sub(r'\s*\[[^\]]*\]\s*$', '', s)     # drop a trailing '[...]' tag
    s = s.replace('$', '').strip()
    return s or f"Contest {cid}"

# DK standings downloads are named 'contest-standings-<id>.csv' (or .zip). Match
# the id loosely so separators can't trip us up: hyphens/underscores/spaces are
# all optional, so 'contest-standings-123', 'conteststandings123', and
# 'contest-standings-123 (1).csv' all resolve to id 123. The .csv/.zip check is
# done separately on the file extension.
_STANDINGS_RE = re.compile(r'contest[-_ ]?standings[-_ ]?(\d+)', re.I)

def discover_contests(wb_path: str):
    """Build the contest list from the DraftKings standings files sitting next to
    the workbook. Each 'contest-standings-<id>.csv' (or .zip) you saved is one
    contest to sim; if there's no file for a contest, it simply isn't simmed.
    Name and entry fee come from dk_lobby.json, payouts from dk_payouts/<id>.json
    — so the Excel 'Contests' sheet isn't needed at all. Returns [] when no
    standings files are present, so the workbook 'Contests' sheet is used instead."""
    base = Path(wb_path).resolve().parent
    found = {}  # id -> relative path (prefer .csv over .zip if both exist)
    # Scan the workbook's folder AND a standings subfolder next to it (named
    # 'standings' or 'Contest Standings', any casing), so the week's downloads
    # can live in one tidy place. Neither scan is recursive: archive old weeks
    # in a sub-subfolder (e.g. Contest Standings\old\) and they're ignored
    # instead of getting simmed again.
    scan_dirs = [base]
    for d in sorted(base.iterdir()):
        if d.is_dir() and d.name.strip().lower() in ("standings", "contest standings"):
            scan_dirs.append(d)
    for d in scan_dirs:
        for f in sorted(d.iterdir()):
            if f.suffix.lower() not in ('.csv', '.zip'):
                continue
            m = _STANDINGS_RE.search(f.name)
            if m:
                cid = m.group(1)
                rel = f.name if d == base else f"{d.name}/{f.name}"
                if cid not in found or (f.suffix.lower() == ".csv"
                                        and found[cid].lower().endswith(".zip")):
                    found[cid] = rel
    if not found:
        return []
    lobby = _load_dk_lobby(base)
    contests = []
    for cid, fname in found.items():
        rec = lobby.get(cid, {})
        fee = rec.get("entry_fee")
        name = rec.get("name")
        if fee is None or not name:
            # Fall back to the scraped payout file: it carries both the entry
            # fee AND the contest name, so a missing/stale/deleted
            # dk_lobby.json can't blank the names on the site.
            pf = base / "dk_payouts" / f"{cid}.json"
            if pf.exists():
                try:
                    with open(pf, "r", encoding="utf-8") as pfh:
                        prec = json.load(pfh)
                    if fee is None:
                        fee = prec.get("entry_fee")
                    if not name:
                        name = prec.get("name")
                except Exception:
                    pass
        if not fee or float(fee) <= 0:
            raise ValueError(
                f"Found {fname} but no entry fee for contest {cid}. "
                f"Refresh the lobby:  python scrape_dk_contests.py")
        contests.append({
            "ContestName": _tidy_contest_name(name, cid),
            "LineupsSheet": fname,        # read_lineups handles .csv and .zip
            "PayoutsSheet": cid,          # load_payouts -> dk_payouts/<id>.json
            "PrefixSheet": None,
            "EntryFee": float(fee),
            "ContestId": cid,
        })
    # De-collide display names (DK runs several identically-named contests per
    # slate, e.g. three "$5 Triple Up"s): identical names would overwrite each
    # other's output files, so repeats get " (2)", " (3)" in contest-id order.
    name_counts = {}
    for c in contests:
        nm = c["ContestName"]
        name_counts[nm] = name_counts.get(nm, 0) + 1
        if name_counts[nm] > 1:
            c["ContestName"] = f"{nm} ({name_counts[nm]})"
    log(f"[discover] {len(contests)} contest(s) from standings files: "
        + ", ".join(c["ContestId"] for c in contests))
    return contests

def read_contests(xl: pd.ExcelFile, wb_path: str):
    """
    If 'Contests' sheet exists, use it. Otherwise, default single contest.
    """
    if "Contests" not in xl.sheet_names:
        return [{
            "ContestName": "Main",
            "LineupsSheet": "Post Contest Sim",
            "PayoutsSheet": "Payouts",
            "PrefixSheet": "PayoutPrefix" if "PayoutPrefix" in xl.sheet_names else None,
            "EntryFee": read_entry_fee_fallback(wb_path),
        }]
    df = pd.read_excel(xl, sheet_name="Contests", engine="openpyxl")
    if df.empty:
        raise ValueError("Contests sheet exists but is empty.")
    cols = {c: safe_str(c).strip().casefold() for c in df.columns}
    inv = {v: k for k, v in cols.items()}
    def get_col(key, required=False):
        k = key.casefold()
        if k not in inv:
            if required:
                raise ValueError(f"Contests sheet missing required column: {key}")
            return None
        return inv[k]
    c_contest = get_col("ContestName", required=True)
    c_lineups = get_col("LineupsSheet", required=True)
    c_payouts = get_col("PayoutsSheet", required=True)
    c_prefix  = get_col("PrefixSheet", required=False)
    c_fee     = get_col("EntryFee", required=True)
    contests = []
    for _, r in df.iterrows():
        name    = safe_str(r[c_contest]) or "Contest"
        lineups = safe_str(r[c_lineups])
        payouts = safe_str(r[c_payouts])
        prefix  = safe_str(r[c_prefix]) if c_prefix else ""
        fee     = to_money(r[c_fee])
        if not lineups or not payouts:
            continue
        if fee <= 0:
            raise ValueError(f"Invalid EntryFee for contest '{name}' in Contests sheet.")
        contests.append({
            "ContestName": name,
            "LineupsSheet": lineups,
            "PayoutsSheet": payouts,
            "PrefixSheet": prefix if prefix else None,
            "EntryFee": float(fee),
        })
    if not contests:
        raise ValueError("No valid contest rows found in Contests sheet.")
    return contests
# -------------------- Copies / keys --------------------
def lineup_key(row6) -> str:
    names = [safe_str(x) for x in row6 if safe_str(x)]
    names.sort()
    return "|".join(names)
def compute_copies_and_keys(fighters_obj_array: np.ndarray):
    n = fighters_obj_array.shape[0]
    keys = [lineup_key(fighters_obj_array[i, :]) for i in range(n)]
    counts = {}
    for k in keys:
        counts[k] = counts.get(k, 0) + 1
    copies = np.array([counts[k] for k in keys], dtype=np.int32)
    return keys, copies
# -------------------- Fight sims --------------------
def load_fight_sims(xl: pd.ExcelFile, fight_ids):
    """Load per-fight sim columns (B=DK F1, C=DK F2). Fail if any fight sheet is missing/empty."""
    S1, S2, N = [], [], []
    missing_or_empty = []
    for fid in fight_ids:
        ok = True
        try:
            df = pd.read_excel(xl, sheet_name=str(fid), engine="openpyxl", usecols="A:C")
            s1 = pd.to_numeric(df.iloc[:, 1], errors="coerce").dropna().to_numpy(np.float32)
            s2 = pd.to_numeric(df.iloc[:, 2], errors="coerce").dropna().to_numpy(np.float32)
            m = int(min(len(s1), len(s2)))
            if m <= 0:
                ok = False
        except Exception:
            ok = False
            s1 = np.zeros(0, np.float32)
            s2 = np.zeros(0, np.float32)
            m = 0
        if not ok:
            missing_or_empty.append(fid)
        S1.append(s1[:m])
        S2.append(s2[:m])
        N.append(m)
    if missing_or_empty:
        raise ValueError(f"Missing or empty sim sheet(s) for FightID(s): {sorted(missing_or_empty)}")
    return S1, S2, N
def build_mats(fighters, fmap, id2idx):
    n = fighters.shape[0]
    F = len(id2idx)
    C1 = np.zeros((n, F), dtype=np.int8)
    C2 = np.zeros((n, F), dtype=np.int8)
    mapped6 = partial = empty = 0
    for i in range(n):
        mapped = 0
        for c in range(6):
            name = norm_name_fighter(fighters[i, c])
            t = fmap.get(name)
            if not t:
                continue  # unmapped => contributes 0
            fid, slot = t
            j = id2idx.get(fid)
            if j is None:
                continue
            if slot == 1:
                C1[i, j] += 1
            else:
                C2[i, j] += 1
            mapped += 1
        if mapped == 6:
            mapped6 += 1
        elif mapped == 0:
            empty += 1
        else:
            partial += 1
    log(f"[map] lineups mapped: 6/6={mapped6:,} | 1–5/6={partial:,} | 0/6={empty:,} (unmapped fighters score 0)")
    return C1, C2
# -------------------- Worker --------------------
def worker_run(idx: int, npz_path: str, iters: int, batch: int, seed: int,
               mem_budget_mb: int = 1024, force_legacy: bool = False):
    rng = np.random.default_rng(seed)
    data = np.load(npz_path, allow_pickle=False)
    # shared fight sampling
    S1_stack = data["S1_stack"]
    S2_stack = data["S2_stack"]
    N = data["N"].astype(np.int64)
    F = int(data["F"])
    # optimal-lineup setup
    opt_fidx = data["opt_fidx"].astype(np.int64)
    opt_slot1 = (data["opt_slot"].astype(np.int64) == 1)
    opt_combos = data["opt_combos"].astype(np.int64)
    opt_sal = data["opt_sal"].astype(np.float64)
    nf_opt = int(data["nf_opt"])
    opt_counts = np.zeros(nf_opt, dtype=np.int64)
    # membership matrix M (nf_opt x C): combo_scores = fs @ M
    if opt_combos.shape[0] and nf_opt:
        M_opt = np.zeros((nf_opt, opt_combos.shape[0]), dtype=np.float32)
        for ci in range(OPT_ROSTER):
            M_opt[opt_combos[:, ci], np.arange(opt_combos.shape[0])] = 1.0
        combo_buf = np.empty((512, opt_combos.shape[0]), dtype=np.float32)
    else:
        M_opt = None
        combo_buf = None
    # contests packed
    K = int(data["K"])
    last_paid = data["last_paid"].astype(np.int64)      # (K,)
    prefix_mat = data["prefix_mat"].astype(np.float64)  # (K, max_rank+1)
    n_list = data["n_list"].astype(np.int64)            # (K,)
    offsets = data["offsets"].astype(np.int64)          # (K+1,)
    C1_concat = data["C1_concat"].astype(np.int8)       # (sum_n, F)
    C2_concat = data["C2_concat"].astype(np.int8)       # (sum_n, F)
    # dedupe structures: entries sharing the exact same fighter columns are one
    # unique lineup; all rank/payout math can run on uniques with copy weights
    U_list = data["U_list"].astype(np.int64)            # (K,) uniques per contest
    u_offsets = data["u_offsets"].astype(np.int64)      # (K+1,)
    uC1_concat = data["uC1_concat"].astype(np.int8)     # (sum_U, F)
    uC2_concat = data["uC2_concat"].astype(np.int8)
    inv_concat = data["inv_concat"].astype(np.int64)    # entry -> unique id, per contest
    # portfolio tracking
    user_map_concat = data["user_map_concat"].astype(np.int64)
    num_users = int(data["num_users"])
    user_total_fees = data["user_total_fees"].astype(np.float64)
    user_contest_fees = data["user_contest_fees"].astype(np.float64)  # (K, num_users)
    # per-contest user maps + precomputed group-reduce structures
    user_map_list = []
    user_perm_list = []     # permutation sorting lineups by user id
    user_starts_list = []   # group starts within the permuted order
    user_ids_list = []      # user id of each group
    for k in range(K):
        a = int(offsets[k]); b = int(offsets[k+1])
        umap = user_map_concat[a:b]
        user_map_list.append(umap)
        perm = np.argsort(umap, kind='stable')
        sorted_umap = umap[perm]
        if len(sorted_umap):
            starts = np.flatnonzero(np.r_[True, sorted_umap[1:] != sorted_umap[:-1]])
            uids = sorted_umap[starts]
        else:
            starts = np.zeros(0, dtype=np.int64)
            uids = np.zeros(0, dtype=np.int64)
        user_perm_list.append(perm)
        user_starts_list.append(starts)
        user_ids_list.append(uids)
    # per-contest transposed matrices (legacy entry columns kept for the
    # fallback path and the startup self-check) + unique-lineup columns
    C1T_list, C2T_list = [], []
    uC1T_list, uC2T_list = [], []
    inv_list, w_list, invperm_list = [], [], []
    for k in range(K):
        a = int(offsets[k]); b = int(offsets[k+1])
        C1T_list.append(C1_concat[a:b].T.astype(np.float32, copy=False))  # (F x n_k)
        C2T_list.append(C2_concat[a:b].T.astype(np.float32, copy=False))
        ua = int(u_offsets[k]); ub = int(u_offsets[k+1])
        uC1T_list.append(uC1_concat[ua:ub].T.astype(np.float32, copy=False))  # (F x U_k)
        uC2T_list.append(uC2_concat[ua:ub].T.astype(np.float32, copy=False))
        inv = inv_concat[a:b]
        inv_list.append(inv)
        w_list.append(np.bincount(inv, minlength=int(U_list[k])).astype(np.int64))
        invperm_list.append(inv[user_perm_list[k]])   # fused gather order for reduceat
    # Fast path applies per contest when it has >=2 unique lineups. Degenerate
    # GEMM shapes (a 1-row micro-batch, or a 1-unique contest) dispatch to GEMV
    # kernels whose summation order differs from the wide GEMM, so those cases
    # always run the legacy path to keep outputs bit-identical.
    fast_ok = [(not force_legacy) and int(U_list[k]) >= 2 for k in range(K)]
    self_checked = [False] * K
    # accumulators (always per-entry; fast batches accumulate per-unique batch
    # sums and expand the length-U vector through inv — identical addends)
    sum_scores = [np.zeros(int(n_list[k]), dtype=np.float64) for k in range(K)]
    sumsq_scores = [np.zeros(int(n_list[k]), dtype=np.float64) for k in range(K)]
    total_payout = [np.zeros(int(n_list[k]), dtype=np.float64) for k in range(K)]
    wins = [np.zeros(int(n_list[k]), dtype=np.float64) for k in range(K)]
    win_total = [np.zeros(int(n_list[k]), dtype=np.float64) for k in range(K)]
    cashes = [np.zeros(int(n_list[k]), dtype=np.float64) for k in range(K)]
    seconds = [np.zeros(int(n_list[k]), dtype=np.float64) for k in range(K)]
    thirds = [np.zeros(int(n_list[k]), dtype=np.float64) for k in range(K)]
    # Portfolio outcome tracking: per-user net profit for each iteration.
    # "all" stays dense over every user; per-contest arrays only hold the users
    # actually IN that contest (everyone else's cell is structurally 0 and the
    # percentile builder excludes all-zero users anyway).
    user_outcomes = np.zeros((iters, num_users), dtype=np.float32)
    contest_outcomes = [np.zeros((iters, len(user_ids_list[k])), dtype=np.float32)
                        for k in range(K)]
    contest_fees_uids = [user_contest_fees[k][user_ids_list[k]] for k in range(K)]
    iter_cursor = 0  # tracks which iteration we're writing to in user_outcomes
    done_total = 0
    while done_total < iters:
        B = min(batch, iters - done_total)
        bytes_budget = int(mem_budget_mb) * (1 << 20)
        # NOTE: this sizing formula is intentionally unchanged. micro_b sets the
        # partial-sum boundaries of every accumulator; changing it would change
        # float64 rounding and break bit-identical output vs prior versions.
        max_n = int(n_list.max()) if len(n_list) else 1
        bytes_per_row = max(1, max_n * 64 + num_users * 16)
        micro_b = max(1, min(B, bytes_budget // bytes_per_row))
        off = 0
        while off < B:
            m = min(micro_b, B - off)
            # sample fights ONCE for m rows
            s1 = np.empty((m, F), dtype=np.float32)
            s2 = np.empty((m, F), dtype=np.float32)
            for f in range(F):
                idxs = rng.integers(0, N[f], size=m, dtype=np.int64)
                s1[:, f] = S1_stack[f, idxs]
                s2[:, f] = S2_stack[f, idxs]
            # ----- optimal-lineup tally (slate-level) -----
            if M_opt is not None:
                fs = np.where(opt_slot1[None, :], s1[:, opt_fidx], s2[:, opt_fidx]).astype(np.float32)
                # Shortcut: if the plain top-6 scorers fit the cap, they ARE
                # the optimal lineup (any other subset scores <=). Only the
                # cap-violating iterations need the brute-force combo argmax.
                top6 = np.argpartition(-fs, OPT_ROSTER - 1, axis=1)[:, :OPT_ROSTER]  # (m, 6)
                fits = opt_sal[top6].sum(axis=1) <= OPT_CAP
                if fits.any():
                    np.add.at(opt_counts, top6[fits].ravel(), 1)
                rem = np.flatnonzero(~fits)
                sub = 512
                for o0 in range(0, len(rem), sub):
                    block = fs[rem[o0:o0+sub]]             # (sm, nf_opt)
                    combo_scores = np.matmul(block, M_opt, out=combo_buf[:block.shape[0]])
                    best = opt_combos[combo_scores.argmax(axis=1)]   # (sm, OPT_ROSTER)
                    np.add.at(opt_counts, best.ravel(), 1)
            # per-user payout accumulator for this micro-batch
            user_payout_batch = np.zeros((m, num_users), dtype=np.float64)
            # evaluate each contest — fully vectorized across the micro-batch
            for k in range(K):
                prefix = prefix_mat[k]
                lp = int(last_paid[k])
                n_k = int(n_list[k])
                U_k = int(U_list[k])
                inv = inv_list[k]
                legacy32 = None
                use_fast = fast_ok[k] and m > 1
                if use_fast:
                    su32 = s1 @ uC1T_list[k]               # (m x U_k) on unique lineups
                    su32 += s2 @ uC2T_list[k]
                    if not self_checked[k]:
                        # One-time self-check: unique-column scores expanded to
                        # entries must be bit-identical to the legacy entry-level
                        # GEMM. If a BLAS/numpy change ever breaks that, this
                        # contest silently drops to the legacy path.
                        legacy32 = s1 @ C1T_list[k] + s2 @ C2T_list[k]
                        if np.array_equal(su32[:, inv].view(np.uint32), legacy32.view(np.uint32)):
                            self_checked[k] = True
                        else:
                            log(f"[warn] dedupe self-check failed for contest {k}; using legacy path")
                            fast_ok[k] = False
                            use_fast = False
                if use_fast:
                    # ---- fast path: rank unique lineups with copy weights ----
                    scu = su32.astype(np.float64)
                    sum_scores[k] += scu.sum(axis=0)[inv]
                    sumsq_scores[k] += np.einsum('ij,ij->j', scu, scu)[inv]
                    # Ranks from one argsort + tie-run boundaries. float32 order
                    # equals float64 order (the cast is monotonic), so ranks and
                    # tie groups are identical to the legacy searchsorted's.
                    negu = -su32
                    order = np.argsort(negu, axis=1)
                    srt = np.take_along_axis(negu, order, axis=1)
                    new_grp = np.empty((m, U_k), dtype=bool)
                    new_grp[:, 0] = True
                    np.not_equal(srt[:, 1:], srt[:, :-1], out=new_grp[:, 1:])
                    posU = np.arange(U_k, dtype=np.int64)
                    start_pos = np.maximum.accumulate(np.where(new_grp, posU, 0), axis=1)
                    end_pos = np.empty((m, U_k), dtype=np.int64)
                    end_pos[:, -1] = U_k
                    em = np.where(new_grp[:, 1:], posU[1:], U_k)
                    np.minimum.accumulate(em[:, ::-1], axis=1, out=end_pos[:, :-1][:, ::-1])
                    # weighted rank interval [Lw, Rw) = entry-level [left, right)
                    w_sorted = w_list[k][order]
                    cw0 = np.empty((m, U_k + 1), dtype=np.int64)
                    cw0[:, 0] = 0
                    np.cumsum(w_sorted, axis=1, out=cw0[:, 1:])
                    Lw_s = np.take_along_axis(cw0, start_pos, axis=1)
                    Rw_s = np.take_along_axis(cw0, end_pos, axis=1)
                    Lw = np.empty((m, U_k), dtype=np.int64)
                    Rw = np.empty((m, U_k), dtype=np.int64)
                    np.put_along_axis(Lw, order, Lw_s, axis=1)
                    np.put_along_axis(Rw, order, Rw_s, axis=1)
                    group_sizes = (Rw - Lw).astype(np.float64)
                    safe_left = np.minimum(Lw, lp)
                    safe_right = np.minimum(Rw, lp)
                    payout = (prefix[safe_right] - prefix[safe_left]) / np.maximum(group_sizes, 1.0)
                    total_payout[k] += payout.sum(axis=0)[inv]
                    cashes[k] += (payout > 0.0).sum(axis=0)[inv]
                    # 1st place: rank interval starting at 0 (ties all count)
                    is_win = Lw == 0
                    wins[k] += is_win.sum(axis=0)[inv]
                    win_total[k] += (payout * is_win).sum(axis=0)[inv]
                    # 2nd/3rd place: the tie groups whose left rank equals the
                    # previous group's right rank (verified identical to the
                    # legacy masked-max computation)
                    rows_i = np.arange(m)
                    wr = Rw[rows_i, np.argmax(is_win, axis=1)]
                    is_second = Lw == wr[:, None]
                    seconds[k] += is_second.sum(axis=0)[inv]
                    has2 = is_second.any(axis=1)
                    r2 = np.where(has2, Rw[rows_i, np.argmax(is_second, axis=1)], -1)
                    thirds[k] += (Lw == r2[:, None]).sum(axis=0)[inv]
                    # per-entry payouts in user-sorted order, gathered straight
                    # from the unique payouts (same values the legacy
                    # payout[:, perm] copy would hold)
                    payout_perm = payout[:, invperm_list[k]]
                else:
                    # ---- legacy path (original engine, kept verbatim) ----
                    if legacy32 is None:
                        legacy32 = s1 @ C1T_list[k] + s2 @ C2T_list[k]
                    sc = legacy32.astype(np.float64)
                    sum_scores[k] += sc.sum(axis=0)
                    sumsq_scores[k] += np.einsum('ij,ij->j', sc, sc)
                    neg = -sc
                    neg_sorted = np.sort(neg, axis=1)          # (m, n_k) ascending = best first
                    row_min = neg_sorted[:, 0].min()
                    row_max = neg_sorted[:, -1].max()
                    span = (row_max - row_min) + 1.0
                    base = (np.arange(m, dtype=np.float64) * span)[:, None]
                    flat_sorted = (neg_sorted + base).ravel()
                    flat_q = (neg + base).ravel()
                    row_start = (np.arange(m, dtype=np.int64) * n_k)[:, None]
                    left = np.searchsorted(flat_sorted, flat_q, side='left').reshape(m, n_k) - row_start
                    right = np.searchsorted(flat_sorted, flat_q, side='right').reshape(m, n_k) - row_start
                    group_sizes = (right - left).astype(np.float64)
                    safe_left = np.minimum(left, lp)
                    safe_right = np.minimum(right, lp)
                    payout = (prefix[safe_right] - prefix[safe_left]) / np.maximum(group_sizes, 1.0)
                    total_payout[k] += payout.sum(axis=0)
                    cashes[k] += (payout > 0.0).sum(axis=0)
                    top = sc.max(axis=1)
                    is_win = sc == top[:, None]
                    wins[k] += is_win.sum(axis=0)
                    win_total[k] += (payout * is_win).sum(axis=0)
                    sc_m = np.where(is_win, -np.inf, sc)
                    val2 = sc_m.max(axis=1)
                    has2 = np.isfinite(val2)
                    is_second = (sc == val2[:, None]) & has2[:, None]
                    seconds[k] += is_second.sum(axis=0)
                    sc_m = np.where(is_second, -np.inf, sc_m)
                    val3 = sc_m.max(axis=1)
                    has3 = np.isfinite(val3)
                    thirds[k] += ((sc == val3[:, None]) & has3[:, None]).sum(axis=0)
                    payout_perm = payout[:, user_perm_list[k]]
                # Portfolio: per-user payouts via grouped reduce (one call per batch)
                starts = user_starts_list[k]
                uids = user_ids_list[k]
                if len(uids):
                    per_user = np.add.reduceat(payout_perm, starts, axis=1)  # (m, n_groups)
                    user_payout_batch[:, uids] += per_user
                    contest_outcomes[k][iter_cursor:iter_cursor + m] = \
                        (per_user - contest_fees_uids[k]).astype(np.float32)
            # Store per-user net profit (payouts - entry fees) for this micro-batch
            user_payout_batch -= user_total_fees  # subtract fees to get net profit
            user_outcomes[iter_cursor:iter_cursor + m] = user_payout_batch.astype(np.float32)
            iter_cursor += m
            off += m
        done_total += B
    # Save user_outcomes to temp files to avoid pipe size limits on Windows
    outcomes_path = npz_path + f".user_outcomes_{idx}.npy"
    np.save(outcomes_path, user_outcomes)
    per_contest_path = npz_path + f".contest_outcomes_{idx}.npz"
    np.savez(per_contest_path, **{f"c{k}": contest_outcomes[k] for k in range(K)})
    return (idx, done_total, sum_scores, sumsq_scores, total_payout, wins, win_total, cashes, seconds, thirds, outcomes_path, per_contest_path, opt_counts)
# -------------------- Pack workbook once --------------------
def pack_npz_multi(wb_path: str, temp_dir: Path):
    xl = pd.ExcelFile(wb_path, engine="openpyxl")
    # Prefer the DraftKings standings files dropped next to the workbook
    # (contest-standings-<id>.csv/zip); fall back to the Excel 'Contests' sheet.
    contests = discover_contests(wb_path) or read_contests(xl, wb_path)
    fmap, fixed_scores, fighter_order, salary_map = read_fighter_map(xl)
    # Build fight_card from fighter_order (preserves DK Fighter Pool sheet order)
    fight_card_map = {}
    fight_card_order = []
    for f in fighter_order:
        fid = f["fight_id"]
        if fid not in fight_card_map:
            fight_card_map[fid] = {"fight_id": fid, "fighter1": None, "fighter2": None,
                                   "fighter1_score": None, "fighter2_score": None}
            fight_card_order.append(fid)
        if f["slot"] == 1:
            fight_card_map[fid]["fighter1"] = f["name"]
            fight_card_map[fid]["fighter1_score"] = fixed_scores.get(f["name"])
        else:
            fight_card_map[fid]["fighter2"] = f["name"]
            fight_card_map[fid]["fighter2_score"] = fixed_scores.get(f["name"])
    fight_card = [fight_card_map[fid] for fid in fight_card_order]
    fights = sorted(set(fid for (fid, _) in fmap.values()))
    id2idx = {fid: i for i, fid in enumerate(fights)}
    # Build fight_fighters map FIRST to determine which fights have fixed scores
    fight_fighters = {}  # fid -> {slot: (name, score_or_None)}
    for name, (fid, slot) in fmap.items():
        if fid not in fight_fighters:
            fight_fighters[fid] = {}
        fight_fighters[fid][slot] = (name, fixed_scores.get(name))
    # Determine which fights are fully fixed (both fighters have scores)
    fixed_fights = set()
    for fid in fights:
        ff = fight_fighters.get(fid, {})
        f1 = ff.get(1)
        f2 = ff.get(2)
        if f1 and f2 and f1[1] is not None and f2[1] is not None:
            fixed_fights.add(fid)
    sim_fights = [fid for fid in fights if fid not in fixed_fights]
    # Only load sim sheets for fights that need them
    if sim_fights:
        sim_S1, sim_S2, sim_N = load_fight_sims(xl, sim_fights)
        sim_map = {fid: i for i, fid in enumerate(sim_fights)}
    # Build final S1/S2/N lists in fight order
    S1_list, S2_list, N_list = [], [], []
    for fid in fights:
        if fid in fixed_fights:
            ff = fight_fighters[fid]
            f1 = ff[1]
            f2 = ff[2]
            S1_list.append(np.array([f1[1]], dtype=np.float32))
            S2_list.append(np.array([f2[1]], dtype=np.float32))
            N_list.append(1)
            log(f"[fixed] fight {fid}: {f1[0]}={f1[1]:.2f}, {f2[0]}={f2[1]:.2f}")
        else:
            si = sim_map[fid]
            S1_list.append(sim_S1[si])
            S2_list.append(sim_S2[si])
            N_list.append(sim_N[si])
    fixed_count = len(fixed_fights)
    if fixed_count:
        log(f"[info] {fixed_count} fight(s) locked with fixed scores, {len(fights) - fixed_count} fight(s) simulated")
    F = len(fights)
    maxN = int(max(N_list)) if N_list else 1
    S1_stack = np.zeros((F, maxN), dtype=np.float32)
    S2_stack = np.zeros((F, maxN), dtype=np.float32)
    N = np.zeros(F, dtype=np.int64)
    for i in range(F):
        Ni = int(N_list[i]); N[i] = Ni
        S1_stack[i, :Ni] = S1_list[i][:Ni]
        S2_stack[i, :Ni] = S2_list[i][:Ni]
    contest_meta = []
    C1_blocks = []
    C2_blocks = []
    n_list = []
    entry_fees = []
    last_paid_list = []
    prefix_list = []
    user_map_blocks = []  # per-contest array mapping lineup_idx -> user_idx
    all_users_set = {}    # username -> user_idx (case-preserved, keyed by casefold)
    # dedupe: unique fighter-column signatures per contest (identical columns =>
    # identical scores every iteration, so the worker ranks uniques w/ weights)
    inv_blocks = []       # per-contest entry -> unique id
    uC1_blocks = []
    uC2_blocks = []
    U_list = []
    for c in contests:
        name = c["ContestName"]
        lineups_ref = c["LineupsSheet"]
        payouts_sheet = c["PayoutsSheet"]
        prefix_sheet = c.get("PrefixSheet", None)
        entry = float(c["EntryFee"])
        payouts_arr, prefix, last_paid = load_payouts(xl, payouts_sheet, prefix_sheet, wb_path)
        prize_pool = float(prefix[last_paid])
        # Read lineups from CSV file or Excel sheet
        fighters, users = read_lineups(xl, lineups_ref, wb_path)
        lineup_keys, copies = compute_copies_and_keys(fighters)
        C1, C2 = build_mats(fighters, fmap, id2idx)
        # Unique lineup signatures, keyed on the exact mapped fighter columns
        # (not name strings): the columns are what determine the score bits.
        sig = np.concatenate([C1, C2], axis=1)
        _, first_idx, inv_k = np.unique(sig, axis=0, return_index=True, return_inverse=True)
        inv_k = inv_k.reshape(-1).astype(np.int64)
        inv_blocks.append(inv_k)
        uC1_blocks.append(C1[first_idx])
        uC2_blocks.append(C2[first_idx])
        U_list.append(len(first_idx))
        # Compute per-lineup total salary from salary_map
        n_k_tmp = fighters.shape[0]
        lineup_salaries = np.zeros(n_k_tmp, dtype=np.float64)
        for i_lu in range(n_k_tmp):
            sal = 0.0
            for c_slot in range(6):
                fname = norm_name_fighter(fighters[i_lu, c_slot])
                if fname and fname in salary_map:
                    sal += salary_map[fname]
            lineup_salaries[i_lu] = sal
        contest_meta.append({
            "Contest": name,
            "LineupsSheet": lineups_ref,
            "PayoutsSheet": payouts_sheet,
            "PrefixSheet": prefix_sheet,
            "EntryFee": entry,
            "PrizePool": prize_pool,
            "fighters": fighters,
            "users": users,
            "lineup_keys": lineup_keys,
            "copies": copies,
            "lineup_salaries": lineup_salaries,
            "n": int(fighters.shape[0]),
            "last_paid": int(last_paid),
            "payouts_array": payouts_arr.tolist(),
        })
        # Build per-lineup user index mapping for portfolio tracking
        n_k = int(fighters.shape[0])
        umap_k = np.zeros(n_k, dtype=np.int64)
        for i_lu, u in enumerate(users):
            ukey = u.casefold() if u else ""
            if ukey not in all_users_set:
                all_users_set[ukey] = (len(all_users_set), u)  # (idx, display_name)
            umap_k[i_lu] = all_users_set[ukey][0]
        user_map_blocks.append(umap_k)
        contest_meta[-1]["uids"] = np.unique(umap_k)  # users in this contest (sorted)
        C1_blocks.append(C1)
        C2_blocks.append(C2)
        n_list.append(n_k)
        entry_fees.append(float(entry))
        last_paid_list.append(int(last_paid))
        prefix_list.append(prefix.astype(np.float64, copy=False))
    # Build user arrays for portfolio tracking
    num_users = len(all_users_set)
    user_display_names = [""] * num_users
    for ukey, (uidx, dname) in all_users_set.items():
        user_display_names[uidx] = dname
    # Per-user total entry fees across all contests + per-contest fees
    user_total_fees = np.zeros(num_users, dtype=np.float64)
    user_contest_fees = np.zeros((len(contest_meta), num_users), dtype=np.float64)
    for k_idx, meta in enumerate(contest_meta):
        entry_k = float(meta["EntryFee"])
        for i_lu in range(len(user_map_blocks[k_idx])):
            uid = user_map_blocks[k_idx][i_lu]
            user_total_fees[uid] += entry_k
            user_contest_fees[k_idx, uid] += entry_k
    user_map_concat = np.concatenate(user_map_blocks).astype(np.int64)
    log(f"[info] portfolio tracking: {num_users:,} unique users across {len(contest_meta)} contests")
    # pack prefix into matrix with padding
    max_rank = max(len(p) for p in prefix_list) - 1
    prefix_mat = np.zeros((len(prefix_list), max_rank + 1), dtype=np.float64)
    for k, p in enumerate(prefix_list):
        L = len(p)
        prefix_mat[k, :L] = p
        if L < max_rank + 1:
            prefix_mat[k, L:] = p[-1]  # keep cumulative flat after end
    # concatenate lineups
    offsets = [0]
    for n in n_list:
        offsets.append(offsets[-1] + n)
    offsets = np.array(offsets, dtype=np.int64)
    C1_concat = np.vstack(C1_blocks).astype(np.int8, copy=False)
    C2_concat = np.vstack(C2_blocks).astype(np.int8, copy=False)
    # dedupe blocks
    u_offsets = [0]
    for u in U_list:
        u_offsets.append(u_offsets[-1] + u)
    u_offsets = np.array(u_offsets, dtype=np.int64)
    uC1_concat = np.vstack(uC1_blocks).astype(np.int8, copy=False)
    uC2_concat = np.vstack(uC2_blocks).astype(np.int8, copy=False)
    inv_concat = np.concatenate(inv_blocks).astype(np.int64)
    total_u = int(u_offsets[-1])
    log(f"[dedupe] {int(offsets[-1]):,} entries -> {total_u:,} unique lineups "
        f"({(int(offsets[-1]) / max(1, total_u)):.2f}x)")
    # -------- Optimal-lineup setup (slate-level) --------
    opt_names = [n for n in salary_map.keys() if n in fmap]
    opt_fidx = np.array([id2idx[fmap[n][0]] for n in opt_names], dtype=np.int64)
    opt_slot = np.array([fmap[n][1] for n in opt_names], dtype=np.int64)
    opt_sal = np.array([salary_map[n] for n in opt_names], dtype=np.float64)
    nf_opt = len(opt_names)
    if nf_opt >= OPT_ROSTER:
        all_combos = np.array(list(itertools.combinations(range(nf_opt), OPT_ROSTER)), dtype=np.int16)
        feasible = opt_sal[all_combos].sum(axis=1) <= OPT_CAP
        opt_combos = all_combos[feasible]
    else:
        opt_combos = np.zeros((0, OPT_ROSTER), dtype=np.int16)
    log(f"[opt] {nf_opt} salaried fighters | {opt_combos.shape[0]:,} cap-feasible lineups")
    npz_path = temp_dir / f"post_bundle_multi_{uuid.uuid4().hex}.npz"
    np.savez_compressed(
        npz_path,
        # shared fights
        S1_stack=S1_stack, S2_stack=S2_stack, N=N,
        F=np.array(F, dtype=np.int64),
        # optimal-lineup
        opt_fidx=opt_fidx, opt_slot=opt_slot, opt_combos=opt_combos,
        opt_sal=opt_sal,
        nf_opt=np.array(nf_opt, dtype=np.int64),
        # contests
        K=np.array(len(contest_meta), dtype=np.int64),
        entry_fees=np.array(entry_fees, dtype=np.float64),
        last_paid=np.array(last_paid_list, dtype=np.int64),
        prefix_mat=prefix_mat,
        n_list=np.array(n_list, dtype=np.int64),
        offsets=offsets,
        C1_concat=C1_concat,
        C2_concat=C2_concat,
        # dedupe
        U_list=np.array(U_list, dtype=np.int64),
        u_offsets=u_offsets,
        uC1_concat=uC1_concat,
        uC2_concat=uC2_concat,
        inv_concat=inv_concat,
        # portfolio tracking
        user_map_concat=user_map_concat,
        num_users=np.array(num_users, dtype=np.int64),
        user_total_fees=user_total_fees,
        user_contest_fees=user_contest_fees,
    )
    return str(npz_path), contest_meta, fight_card, user_display_names, salary_map, opt_names
# -------------------- Main --------------------
def ask_int(prompt: str, default: int, min_val: int = 1) -> int:
    while True:
        s = input(f"{prompt} [{default}]: ").strip()
        if not s:
            return default
        try:
            v = int(s)
            if v >= min_val:
                return v
        except Exception:
            pass
        print(f"Please enter an integer >= {min_val}.")
def main():
    global _already_paused
    try:
        os.chdir(Path(__file__).resolve().parent)
    except Exception:
        pass
    DEFAULT_WB = "Post Contest Sim.xlsm"
    DEFAULT_ITERS = 200000
    # Suggest roughly the physical core count (cpu_count reports logical/SMT).
    DEFAULT_WORKERS = max(4, min(16, (os.cpu_count() or 8) // 2))
    DEFAULT_BATCH = 32768
    DEFAULT_STEP = 100000
    ap = argparse.ArgumentParser(add_help=False)
    ap.add_argument("--workbook")
    ap.add_argument("--iters", type=int)
    ap.add_argument("--workers", type=int)
    ap.add_argument("--batch", type=int)
    ap.add_argument("--progress_step", type=int)
    ap.add_argument("--seed", type=int)
    ap.add_argument("--out")
    ap.add_argument("--legacy", action="store_true",
                    help="force the original ranking engine (slow path) everywhere")
    args, _ = ap.parse_known_args()
    wb = args.workbook or DEFAULT_WB
    if not Path(wb).exists():
        s = input(f"Workbook not found ({wb}). Enter full path or press Enter to abort: ").strip()
        if s:
            wb = s
    if not Path(wb).exists():
        print(f"Workbook not found: {wb}")
        try: input("Press Enter to close...")
        except Exception: pass
        _already_paused = True
        return
    iters = args.iters if (args.iters and args.iters > 0) else ask_int("Number of iterations", DEFAULT_ITERS)
    workers = args.workers if (args.workers and args.workers > 0) else ask_int(
        f"Number of worker processes (your CPU suggests {DEFAULT_WORKERS})", DEFAULT_WORKERS)
    if args.batch and args.batch > 0:
        batch = args.batch
    else:
        print("Batch size options: 8192, 16384, 32768, 65536, 131072")
        batch = ask_int("Batch size", DEFAULT_BATCH)
    step = args.progress_step if (args.progress_step and args.progress_step > 0) else ask_int("Progress step (iters per update)", DEFAULT_STEP)
    if args.seed is None:
        seed = int.from_bytes(os.urandom(8), "little") & 0x7FFFFFFFFFFFFFFF
        log(f"[info] Using RANDOM seed: {seed}")
    else:
        seed = int(args.seed)
        log(f"[info] Using FIXED seed: {seed}")
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = args.out or f"PostContest_Results_Multi_{iters}_{seed}_{ts}.csv"
    out_dir = str(Path(out_path).resolve().parent)
    base_stem = Path(out_path).stem
    t0 = time.time()
    with tempfile.TemporaryDirectory() as td:
        bundle, contest_meta, fight_card, user_display_names, salary_map, opt_names = pack_npz_multi(wb, Path(td))
        opt_counts_total = np.zeros(len(opt_names), dtype=np.int64)
        K = len(contest_meta)
        log(f"[info] contests={K} | workers={workers} | batch={batch}")
        for c in contest_meta:
            log(f"  - {c['Contest']}: lineups={c['n']:,} | entry=${c['EntryFee']:.2f} | prize_pool=${c['PrizePool']:.2f} | last_paid={c['last_paid']}")
        log("[info] using vectorized batch payout engine (flat searchsorted)")
        # Build chunk plan.
        # v1 made one chunk per progress_step, so with step >= iters/workers
        # some workers never received any work. Chunk size is now capped at
        # ceil(iters/workers) so every worker stays busy; progress still
        # reports at least once per completed chunk.
        chunk_size = max(1, min(step, math.ceil(iters / max(1, workers))))
        chunks = []
        remaining = iters
        while remaining > 0:
            c = chunk_size if remaining > chunk_size else remaining
            chunks.append(c)
            remaining -= c
        rng = np.random.default_rng(seed)
        child_seeds = rng.integers(0, 2**63 - 1, size=len(chunks), dtype=np.int64)
        # per contest accumulators
        sum_scores   = [np.zeros(c["n"], dtype=np.float64) for c in contest_meta]
        sumsq_scores = [np.zeros(c["n"], dtype=np.float64) for c in contest_meta]
        total_payout = [np.zeros(c["n"], dtype=np.float64) for c in contest_meta]
        wins         = [np.zeros(c["n"], dtype=np.float64) for c in contest_meta]
        win_total    = [np.zeros(c["n"], dtype=np.float64) for c in contest_meta]
        cashes       = [np.zeros(c["n"], dtype=np.float64) for c in contest_meta]
        seconds      = [np.zeros(c["n"], dtype=np.float64) for c in contest_meta]
        thirds       = [np.zeros(c["n"], dtype=np.float64) for c in contest_meta]
        done_iters = 0
        all_user_outcomes = []  # collect per-worker user outcome arrays
        contest_npzs = []       # per-worker npz handles of ragged per-contest outcomes
        # Pin BLAS to one thread per worker process: the workers ARE the
        # parallelism, and oversubscribed BLAS threads fight each other. Thread
        # count doesn't change GEMM results here (the k-reduction per output
        # element is sequential either way). Users' own env settings win.
        for _v in ("OMP_NUM_THREADS", "OPENBLAS_NUM_THREADS", "MKL_NUM_THREADS",
                   "NUMEXPR_NUM_THREADS"):
            os.environ.setdefault(_v, "1")
        with ProcessPoolExecutor(max_workers=workers) as ex:
            futs = [ex.submit(worker_run, i, bundle, int(chunks[i]), int(batch), int(child_seeds[i]),
                              force_legacy=bool(args.legacy))
                    for i in range(len(chunks))]
            # Collect in submission order (NOT completion order) so float64
            # accumulation order is deterministic: same seed + same settings
            # now reproduce byte-identical outputs run to run.
            for fut in futs:
                (idx, its, s_list, ss_list, tp_list, w_list, wt_list, c_list, sec_list, thi_list, outcomes_path, per_contest_path, opt_c) = fut.result()
                if opt_c is not None and len(opt_c) == len(opt_counts_total):
                    opt_counts_total += opt_c
                for k in range(K):
                    sum_scores[k]   += s_list[k]
                    sumsq_scores[k] += ss_list[k]
                    total_payout[k] += tp_list[k]
                    wins[k]         += w_list[k]
                    win_total[k]    += wt_list[k]
                    cashes[k]       += c_list[k]
                    seconds[k]      += sec_list[k]
                    thirds[k]       += thi_list[k]
                all_user_outcomes.append(np.load(outcomes_path))
                os.remove(outcomes_path)
                contest_npzs.append(np.load(per_contest_path))  # lazy npz; closed after percentiles
                done_iters += its
                rate = done_iters / max(1e-9, (time.time() - t0))
                log(f"[progress] {done_iters:,}/{iters:,} ({done_iters/iters:,.1%}) | {rate:,.0f} it/s")
        elapsed = time.time() - t0
        log(f"[timing] total wall: {elapsed:,.2f}s | iters/sec: {iters/elapsed:,.0f}")
        # Optimal% per fighter (slate-level): fraction of iterations a fighter is in the optimal lineup
        optimal_map = {opt_names[i]: round(float(opt_counts_total[i]) / max(1, iters) * 100.0, 2)
                       for i in range(len(opt_names))}
        if optimal_map:
            log("[opt] top optimal%: " + ", ".join(
                f"{n}={optimal_map[n]:.1f}%" for n in sorted(optimal_map, key=optimal_map.get, reverse=True)[:5]))
        # Write output per contest (one CSV each)
        for k, meta in enumerate(contest_meta):
            entry = float(meta["EntryFee"])
            n = int(meta["n"])
            mean = sum_scores[k] / max(1, iters)
            var  = (sumsq_scores[k] / max(1, iters)) - (mean * mean)
            var[var < 0] = 0.0
            sd   = np.sqrt(var * iters / max(1, iters - 1)) if iters > 1 else np.zeros_like(var)
            p99  = mean + Z99 * sd
            EV   = total_payout[k] / max(1, iters)
            NetEV = EV - entry
            ROI   = np.where(entry > 0, NetEV / entry * 100.0, 0.0)
            WinPct = wins[k] / max(1, iters) * 100.0
            SecondPct = seconds[k] / max(1, iters) * 100.0
            ThirdPct = thirds[k] / max(1, iters) * 100.0
            CashPct = cashes[k] / max(1, iters) * 100.0
            AvgWinPayout = np.where(wins[k] > 0, win_total[k] / np.maximum(wins[k], 1.0), 0.0)
            fighters = meta["fighters"]
            users    = meta["users"]
            keys     = meta["lineup_keys"]
            copies   = meta["copies"]
            salaries = meta["lineup_salaries"]
            # Column-wise build (same values/dtypes as the old per-row loop,
            # which spent ~15s in Python for large slates)
            df_k = pd.DataFrame({
                "Contest": [meta["Contest"]] * n,
                "Row": np.arange(1, n + 1, dtype=np.int64),
                "Username": list(users),
                "LineupKey": list(keys),
                "Copies": copies.astype(np.int64),
                "F1": [safe_str(x) for x in fighters[:, 0]],
                "F2": [safe_str(x) for x in fighters[:, 1]],
                "F3": [safe_str(x) for x in fighters[:, 2]],
                "F4": [safe_str(x) for x in fighters[:, 3]],
                "F5": [safe_str(x) for x in fighters[:, 4]],
                "F6": [safe_str(x) for x in fighters[:, 5]],
                "EntryFee": np.full(n, float(entry), dtype=np.float64),
                "AvgWinPayout": AvgWinPayout.astype(np.float64, copy=False),
                "EV": EV.astype(np.float64, copy=False),
                "NetEV": NetEV.astype(np.float64, copy=False),
                "ROI%": ROI.astype(np.float64, copy=False),
                "WinPct": WinPct.astype(np.float64, copy=False),
                "SecondPct": SecondPct.astype(np.float64, copy=False),
                "ThirdPct": ThirdPct.astype(np.float64, copy=False),
                "CashPct": CashPct.astype(np.float64, copy=False),
                "MeanScore": mean.astype(np.float64, copy=False),
                "SDScore": sd.astype(np.float64, copy=False),
                "P99Score": p99.astype(np.float64, copy=False),
                "TotalPayout": total_payout[k].astype(np.float64, copy=False),
                "TotalSalary": salaries.astype(np.float64, copy=False),
            })
            cname = safe_filename(meta["Contest"])
            per_path = os.path.join(out_dir, f"{base_stem}_{cname}.csv")
            df_k.to_csv(per_path, index=False, encoding="utf-8")
            log(f"[done] wrote {per_path}")
            # Write companion meta JSON with fight_card and payouts for What If feature
            meta_json_path = os.path.join(out_dir, f"{base_stem}_{cname}_meta.json")
            meta_json = {
                "fight_card": fight_card,
                "payouts": meta["payouts_array"],
                "entry_fee": float(entry),
                "salary_map": salary_map,
                "optimal_map": optimal_map,
            }
            with open(meta_json_path, 'w', encoding='utf-8') as mf:
                json.dump(meta_json, mf, indent=2)
            log(f"[done] wrote {meta_json_path}")
            # sanity check per contest
            prize_pool = float(meta["PrizePool"])
            avg_ev = float(EV.mean()) if n else 0.0
            expected_ev = float(prize_pool / n) if n else 0.0
            total_paid_per_contest = float(total_payout[k].sum() / iters) if iters else 0.0
            log(f"[check:{meta['Contest']}] prize_pool=${prize_pool:,.2f} | entries={n:,} | "
                f"avg EV=${avg_ev:.4f} (expected ${expected_ev:.4f}) | "
                f"avg total paid=${total_paid_per_contest:,.2f}")
            if abs(avg_ev - expected_ev) > 1e-4 or abs(total_paid_per_contest - prize_pool) > 1e-2:
                log(f"[warn:{meta['Contest']}] EV/payout conservation check failed; verify payout prefix/tie-split logic and inputs.")
        # Compute and write portfolio percentile distributions.
        # All sorting is done as bulk column sorts up front (one np.sort per
        # matrix instead of one per user) — identical order statistics, just
        # ~10-20x faster than the old per-user Python loop.
        combined_outcomes = np.vstack(all_user_outcomes)  # (total_iters, num_users)
        del all_user_outcomes  # free memory
        num_users = combined_outcomes.shape[1]
        sorted_all = np.sort(np.ascontiguousarray(combined_outcomes.T), axis=1)  # (num_users, total_iters)
        del combined_outcomes
        percentile_points = [1, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70, 75, 80, 85, 90, 95, 99]
        contest_names = [meta["Contest"] for meta in contest_meta]
        # Per-contest ragged matrices: columns are only the users IN the
        # contest (worker's user_ids_list order = sorted global user ids).
        contest_sorted = []   # (n_uids_k, total_iters) sorted per user
        contest_nz = []       # per-column any-nonzero mask
        contest_uid_pos = []  # global user id -> column position
        for k in range(K):
            pieces = [z[f"c{k}"] for z in contest_npzs]
            mat = np.vstack(pieces)  # (total_iters, n_uids_k)
            del pieces
            uids_k = contest_meta[k]["uids"]
            contest_nz.append((mat != 0).any(axis=0) if mat.size else
                              np.zeros(len(uids_k), dtype=bool))
            contest_sorted.append(np.sort(np.ascontiguousarray(mat.T), axis=1))
            contest_uid_pos.append({int(u): i for i, u in enumerate(uids_k)})
            del mat
        for z in contest_npzs:
            z.close()

        def compute_pctiles(sorted_data):
            pctiles = {"min": round(float(sorted_data[0]), 2), "max": round(float(sorted_data[-1]), 2)}
            for p in percentile_points:
                idx_p = int(np.floor(p / 100.0 * len(sorted_data)))
                idx_p = min(idx_p, len(sorted_data) - 1)
                pctiles[f"p{p}"] = round(float(sorted_data[idx_p]), 2)
            return pctiles

        portfolio_percentiles = {"contests": contest_names, "users": {}}
        for u_idx in range(num_users):
            uname = user_display_names[u_idx]
            if not uname:
                continue
            user_entry = {"all": compute_pctiles(sorted_all[u_idx]), "by_contest": {}}
            # Per-contest percentiles (only if the user has entries there —
            # same any-nonzero rule as before)
            for k_idx, cname in enumerate(contest_names):
                pos = contest_uid_pos[k_idx].get(u_idx)
                if pos is not None and contest_nz[k_idx][pos]:
                    user_entry["by_contest"][cname] = compute_pctiles(contest_sorted[k_idx][pos])
            portfolio_percentiles["users"][uname] = user_entry
        del sorted_all, contest_sorted  # free memory
        pct_path = os.path.join(out_dir, f"{base_stem}_portfolio_percentiles.json")
        with open(pct_path, 'w', encoding='utf-8') as pf:
            json.dump(portfolio_percentiles, pf, indent=2)
        log(f"[done] wrote portfolio percentiles: {pct_path} ({num_users} users, {K} contests)")
    try:
        input("Press Enter to close...")
    except Exception:
        pass
    _already_paused = True
# ---- safe entrypoint for double-click on Windows ----
def _safe_main():
    try:
        import multiprocessing as mp
        try:
            mp.set_start_method("spawn", force=True)
        except Exception:
            pass
        mp.freeze_support()
    except Exception:
        pass
    try:
        main()
    except Exception as e:
        print("FATAL:", e)
        import traceback
        print(traceback.format_exc())
        try:
            input("Press Enter to close...")
        except Exception:
            pass
if __name__ == "__main__":
    _safe_main()

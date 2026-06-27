"""
Scrape actual fight stats from ufcstats.com and fill in the Actual Scores
columns in the 'Projections vs Actuals.xlsx' workbook.

Only fills rows where the Actual Scores columns are empty.

Usage:
    python scrape_actuals.py                              # default workbook
    python scrape_actuals.py "Projections vs Actuals.xlsx"
"""
from __future__ import annotations

import hashlib
import re
import sys
import time

import openpyxl
import requests

BASE = "http://www.ufcstats.com"
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}

# Shared session so the anti-bot clearance cookie persists across requests.
_SESSION = requests.Session()
_SESSION.headers.update(UA)

# DraftKings scoring multipliers (must match sim/config.py)
DK_SS = 0.2       # sig strike portion (sig-only)
DK_STR = 0.2      # regular strike portion (all strikes including sig)
DK_TD = 5.0
DK_KD = 10.0
DK_CTRL = 0.03    # per second
DK_REV = 5.0

# Finish bonuses
BONUS_DEC = 30.0
BONUS_BY_ROUND = {1: 90.0, 2: 70.0, 3: 45.0, 4: 40.0, 5: 40.0}
BONUS_QWB = 25.0  # quick win bonus (under 60 sec)


def _is_challenge(html: str) -> bool:
    """Detect ufcstats.com's JS proof-of-work anti-bot interstitial."""
    return "Checking your browser" in html or ("/__c" in html and "nonce=" in html)


def _solve_challenge(html: str) -> bool:
    """Solve the proof-of-work interstitial and register clearance on the session.

    The page asks for the smallest n such that sha256(nonce + ':' + n) starts with
    a run of '0' hex chars, then POSTs {nonce, n} to /__c to obtain a cookie.
    Returns True if a challenge was detected and a solution submitted.
    """
    nonce_m = re.search(r'nonce="([^"]+)"', html)
    target_m = re.search(r"target=new Array\((\d+)\+1\)", html)
    if not nonce_m or not target_m:
        return False

    nonce = nonce_m.group(1)
    prefix = "0" * int(target_m.group(1))
    n = 0
    while not hashlib.sha256(f"{nonce}:{n}".encode()).hexdigest().startswith(prefix):
        n += 1
    _SESSION.post(f"{BASE}/__c", data={"nonce": nonce, "n": n}, timeout=15)
    return True


def _get(url: str) -> str:
    """Fetch a URL with retries, transparently clearing the anti-bot challenge."""
    for attempt in range(4):
        try:
            r = _SESSION.get(url, timeout=15)
            r.raise_for_status()
            html = r.text
            # If we hit the proof-of-work wall, solve it and re-fetch once.
            if _is_challenge(html) and _solve_challenge(html):
                r = _SESSION.get(url, timeout=15)
                r.raise_for_status()
                html = r.text
            return html
        except requests.RequestException:
            if attempt < 3:
                time.sleep(2 ** (attempt + 1))
            else:
                raise


def _parse_landed(text: str) -> int:
    """Parse '42 of 73' → 42, or plain int."""
    text = text.strip()
    m = re.match(r"(\d+)\s+of\s+\d+", text)
    if m:
        return int(m.group(1))
    return int(text) if text.isdigit() else 0


def _parse_ctrl(text: str) -> int:
    """Parse '2:04' → 124 seconds, or '--' → 0."""
    text = text.strip()
    if text == "--" or not text:
        return 0
    m = re.match(r"(\d+):(\d+)", text)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    return 0


def _parse_time(text: str) -> int:
    """Parse fight time '4:32' → 272 seconds."""
    text = text.strip()
    m = re.match(r"(\d+):(\d+)", text)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    return 0


def fetch_events() -> list[dict]:
    """Fetch completed events list: [{name, date_str, url}, ...]."""
    html = _get(f"{BASE}/statistics/events/completed")
    events = []
    for m in re.finditer(
        r'event-details/([a-f0-9]+)"[^>]*>\s*([^<]+)</a>\s*'
        r'<span[^>]*>\s*([^<]+)</span>',
        html, re.DOTALL,
    ):
        eid, name, date_str = m.group(1), m.group(2).strip(), m.group(3).strip()
        url = f"{BASE}/event-details/{eid}"
        events.append({"name": name, "date_str": date_str, "url": url})
    return events


def fetch_event_fights(event_url: str) -> list[dict]:
    """Fetch fights from an event page: [{url, fighters: [name1, name2]}, ...]."""
    html = _get(event_url)
    fights = []
    for m in re.finditer(
        r'<tr[^>]*data-link="([^"]+)"[^>]*>(.*?)</tr>', html, re.DOTALL
    ):
        fight_url, content = m.group(1), m.group(2)
        names = [
            n.strip()
            for n in re.findall(r"fighter-details/[^>]+>([^<]+)<", content)
        ]
        if len(names) >= 2:
            # Normalize URL to use base domain
            fid = re.search(r"fight-details/([a-f0-9]+)", fight_url)
            if fid:
                fights.append({
                    "url": f"{BASE}/fight-details/{fid.group(1)}",
                    "fighters": names,
                })
    return fights


def fetch_fight_stats(fight_url: str) -> dict | None:
    """Fetch stats for a single fight.

    Returns dict with keys for each fighter:
      {fighter_name: {ss, total_str, kd, td, rev, ctrl_sec, sub_att,
                      is_winner, method, round, time_sec}, ...}
    """
    html = _get(fight_url)

    # --- Parse winner/loser (W/L/D, plus NC for No Contest) ---
    persons = re.findall(
        r'b-fight-details__person-status[^>]*>\s*(W|L|D|NC)\s*</i>.*?'
        r'b-fight-details__person-link[^>]*>([^<]+)<',
        html, re.DOTALL,
    )
    if len(persons) < 2:
        return None

    result_map = {}
    for status, name in persons:
        result_map[name.strip()] = status.strip()

    # --- Parse method, round, time ---
    method_m = re.search(
        r'Method:.*?font-style:\s*normal[^>]*>\s*([^<]+)<', html, re.DOTALL
    )
    method_str = method_m.group(1).strip() if method_m else ""

    round_m = re.search(r'Round:\s*</i>\s*(\d+)', html, re.DOTALL)
    fight_round = int(round_m.group(1)) if round_m else 0

    time_m = re.search(r'Time:\s*</i>\s*([\d:]+)', html, re.DOTALL)
    time_sec = _parse_time(time_m.group(1)) if time_m else 0

    # Total fight time in seconds: (round-1)*5min + time in final round
    total_time_sec = (fight_round - 1) * 300 + time_sec if fight_round > 0 else 0

    is_decision = "decision" in method_str.lower()

    # --- Parse totals table ---
    tables = html.split("<table")
    if len(tables) < 2:
        return None
    totals_html = tables[1].split("</table>")[0]
    vals = re.findall(
        r'b-fight-details__table-text[^>]*>\s*(.*?)\s*</p>',
        totals_html, re.DOTALL,
    )
    vals = [re.sub(r"<[^>]+>", "", v).strip() for v in vals]

    if len(vals) < 20:
        return None

    # Columns (paired: fighter1, fighter2):
    # 0-1: names, 2-3: KD, 4-5: sig str, 6-7: sig str %,
    # 8-9: total str, 10-11: TD, 12-13: TD %, 14-15: sub att,
    # 16-17: rev, 18-19: ctrl
    names = [vals[0].strip(), vals[1].strip()]

    stats = {}
    for i, name in enumerate(names):
        ss = _parse_landed(vals[4 + i])
        total_str = _parse_landed(vals[8 + i])
        kd = _parse_landed(vals[2 + i])
        td = _parse_landed(vals[10 + i])
        sub_att = _parse_landed(vals[14 + i])
        rev = _parse_landed(vals[16 + i])
        ctrl_sec = _parse_ctrl(vals[18 + i])
        is_winner = result_map.get(name, "") == "W"

        # Compute DK scoring components
        ss_pts = ss * DK_SS
        str_pts = total_str * DK_STR
        td_pts = td * DK_TD
        kd_pts = kd * DK_KD
        ctrl_pts = ctrl_sec * DK_CTRL
        rev_pts = rev * DK_REV

        bonus = 0.0
        if is_winner:
            if is_decision:
                bonus = BONUS_DEC
            else:
                bonus = BONUS_BY_ROUND.get(fight_round, 40.0)
                if total_time_sec <= 60:
                    bonus += BONUS_QWB

        total_pts = ss_pts + str_pts + td_pts + kd_pts + ctrl_pts + rev_pts + bonus

        stats[name] = {
            "ss": ss,
            "total_str": total_str,
            "kd": kd,
            "td": td,
            "sub_att": sub_att,
            "rev": rev,
            "ctrl_sec": ctrl_sec,
            "is_winner": is_winner,
            "method": method_str,
            "round": fight_round,
            "time_sec": time_sec,
            "total_time_sec": total_time_sec,
            # DK point components
            "ss_pts": round(ss_pts, 2),
            "str_pts": round(str_pts, 2),
            "td_pts": round(td_pts, 2),
            "kd_pts": round(kd_pts, 2),
            "ctrl_pts": round(ctrl_pts, 2),
            "rev_pts": round(rev_pts, 2),
            "bonus_pts": round(bonus, 2),
            "total_pts": round(total_pts, 2),
        }

    return stats


def _normalize(name: str) -> str:
    """Normalize a fighter name for matching."""
    return re.sub(r"[^a-z]", "", name.lower())


def _name_tokens(name: str) -> frozenset:
    """Split a name into a set of lowercase word tokens."""
    return frozenset(re.findall(r"[a-z]+", name.lower()))


def _find_stats(fighter: str, stats_cache: dict) -> dict | None:
    """Match a spreadsheet fighter name to scraped stats.

    Tolerates word-order swaps (e.g. 'Yadong Song' vs 'Song Yadong') and dropped
    given-name tokens (e.g. 'Su Sumudaerji' vs 'Sumudaerji') that are common for
    fighters ufcstats.com lists in family-name-first order.
    """
    # 1) exact match on letters-only concatenation
    target = _normalize(fighter)
    for name, st in stats_cache.items():
        if _normalize(name) == target:
            return st
    # 2) identical set of name tokens regardless of order
    qt = _name_tokens(fighter)
    same = [st for name, st in stats_cache.items() if _name_tokens(name) == qt]
    if len(same) == 1:
        return same[0]
    # 3) one name's tokens are a subset of the other's (only if unambiguous)
    subset = [
        st for name, st in stats_cache.items()
        if _name_tokens(name) <= qt or qt <= _name_tokens(name)
    ]
    if len(subset) == 1:
        return subset[0]
    return None


def _find_event_for_date(events: list[dict], date_str: str) -> dict | None:
    """Find an event matching a date string like '3/14/2026' or 'March 14, 2026'."""
    import datetime

    # Parse the target date from various formats
    target = None
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y"):
        try:
            target = datetime.datetime.strptime(date_str.strip(), fmt).date()
            break
        except ValueError:
            continue

    if target is None:
        return None

    for ev in events:
        try:
            ev_date = datetime.datetime.strptime(ev["date_str"].strip(), "%B %d, %Y").date()
            if ev_date == target:
                return ev
        except ValueError:
            continue

    return None


def fill_actuals(xlsx_path: str = "Projections vs Actuals.xlsx"):
    """Read the workbook, find rows with empty actuals, scrape and fill."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Row 1 has section headers ("Predicted Scores" / "Actual Scores")
    # Row 2 has column headers; data starts at row 3
    headers = [cell.value for cell in ws[2]]

    # Find column indices for Date and Fighter (unique in row 2)
    def col_idx(name):
        for i, h in enumerate(headers):
            if h and h.strip() == name:
                return i
        return None

    date_col = col_idx("Date")
    fighter_col = col_idx("Fighter")

    # Actual columns are identified by the "Actual Scores" label in row 1
    # Find the column where "Actual Scores" starts
    act_start = None
    for i, cell in enumerate(ws[1]):
        if cell.value and "Actual" in str(cell.value):
            act_start = i
            break

    if act_start is None:
        print("  ERROR: 'Actual Scores' section not found in row 1.")
        return

    # Map actual column names to their 0-based indices
    # Expected order: SS Pts, Non-Sig Pts, TD Pts, KD Pts, Ctrl Pts, Rev Pts, Bonus Pts, Total Pts
    act_ss = act_start
    act_nonsig = act_start + 1
    act_td = act_start + 2
    act_kd = act_start + 3
    act_ctrl = act_start + 4
    act_rev = act_start + 5
    act_bonus = act_start + 6
    act_total = act_start + 7

    # Collect rows that need filling (date → set of fighter names)
    rows_to_fill = []
    dates_needed = set()
    for row_idx in range(3, ws.max_row + 1):
        fighter = ws.cell(row=row_idx, column=fighter_col + 1).value if fighter_col is not None else None
        date = ws.cell(row=row_idx, column=date_col + 1).value if date_col is not None else None
        actual_total = ws.cell(row=row_idx, column=act_total + 1).value

        if fighter and (actual_total is None or str(actual_total).strip() == ""):
            rows_to_fill.append((row_idx, date, fighter))
            if date:
                dates_needed.add(date)

    if not rows_to_fill:
        print("  All actual scores already filled!")
        wb.close()
        return

    print(f"  Found {len(rows_to_fill)} rows needing actual scores.")

    # Convert dates to date objects for matching
    import datetime
    date_objects = set()
    for d in dates_needed:
        if isinstance(d, datetime.datetime):
            date_objects.add(d.date())
        elif isinstance(d, datetime.date):
            date_objects.add(d)
        else:
            # Try parsing string dates
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y"):
                try:
                    date_objects.add(datetime.datetime.strptime(str(d).strip(), fmt).date())
                    break
                except ValueError:
                    continue

    print(f"  Dates to fetch: {date_objects}")

    # Fetch events list
    print("  Fetching events list from ufcstats.com ...")
    events = fetch_events()
    print(f"  Found {len(events)} events.")

    # For each date, find the event and get all fight stats
    stats_cache = {}  # scraped fighter name → stats dict
    for target_date in date_objects:
        # Match event by parsed date
        ev = None
        for e in events:
            try:
                ev_date = datetime.datetime.strptime(e["date_str"].strip(), "%B %d, %Y").date()
                if ev_date == target_date:
                    ev = e
                    break
            except ValueError:
                continue

        if ev is None:
            print(f"  WARNING: No event found for date '{target_date}'")
            continue

        print(f"  Fetching fights for: {ev['name']} ({ev['date_str']}) ...")
        fights = fetch_event_fights(ev["url"])
        print(f"    {len(fights)} fights found. Fetching stats ...")

        for fi, fight in enumerate(fights):
            print(f"    [{fi+1}/{len(fights)}] {' vs '.join(fight['fighters'])} ...", end="")
            fight_stats = fetch_fight_stats(fight["url"])
            if fight_stats:
                for name, st in fight_stats.items():
                    stats_cache[name] = st
                print(" OK")
            else:
                print(" FAILED")
            time.sleep(0.3)  # polite rate limiting

    # Fill in the actual columns
    # SS Pts = sig_strikes * 0.2, Non-Sig Pts = total_strikes * 0.2
    # (sig strikes are a subset of total strikes; both scored independently)
    filled = 0
    for row_idx, date, fighter in rows_to_fill:
        st = _find_stats(fighter, stats_cache)
        if st is None:
            print(f"  WARNING: No stats found for '{fighter}'")
            continue

        ws.cell(row=row_idx, column=act_ss + 1, value=st["ss_pts"])
        ws.cell(row=row_idx, column=act_nonsig + 1, value=st["str_pts"])
        ws.cell(row=row_idx, column=act_td + 1, value=st["td_pts"])
        ws.cell(row=row_idx, column=act_kd + 1, value=st["kd_pts"])
        ws.cell(row=row_idx, column=act_ctrl + 1, value=st["ctrl_pts"])
        ws.cell(row=row_idx, column=act_rev + 1, value=st["rev_pts"])
        ws.cell(row=row_idx, column=act_bonus + 1, value=st["bonus_pts"])
        ws.cell(row=row_idx, column=act_total + 1, value=st["total_pts"])
        filled += 1

    wb.save(xlsx_path)
    print(f"\n  Done! Filled {filled}/{len(rows_to_fill)} rows in {xlsx_path}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "Projections vs Actuals.xlsx"
    fill_actuals(path)
